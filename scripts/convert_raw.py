#!/usr/bin/env python3
"""RAW_new/ 원본 파일(협회 엑셀·CSV 혼합) → raw/*.csv 표준 헤더 변환.

사용법:
    pip install openpyxl
    python scripts/convert_raw.py --year 2026

여러 연도의 원본이 RAW_new/에 함께 있으면 --year로 대상 연도를 지정한다. 파일명에
연도가 없는 소스(CAK는 '25년도', KICA는 수집 시각)까지 고려해 후보가 여러 개면
어떤 파일을 골라야 할지 물어보고 중단한다.

원본 → 표준 raw 매핑 (파일 명세: docs/PRD.md §3):
    CAK_*.xlsx (종합건설, 5시트)        → raw/general_construction.csv
    전문건설업*_업종.xlsx (12시트)       → raw/specialty_construction.csv
    2025년 기계설비*.xlsx (1시트)        → raw/mechanical_gas.csv
    CCE_*.xlsx (소방설비)               → raw/fire_protection.csv
    KICA_rank_*.csv (정보통신, CP949)   → raw/ict_communication.csv
    전기설비는 크롤링 산출물이므로 이 스크립트가 다루지 않는다 (scripts/crawl_keca.py).

주의: 종합건설·전문건설·기계설비 원본의 '연번'은 가나다순 일련번호이며 순위가 아닌
경우가 있다. 순위 확정은 normalize.py의 rank_policy가 담당한다 (PRD §6.4).
"""
import argparse
import csv
import glob
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
RAW_NEW = ROOT / "RAW_new"
RAW_OUT = ROOT / "raw"

YEAR = None   # --year로 지정. 후보가 여러 개일 때 연도로 좁히는 데 쓴다.

HEADER = ["대분류", "중분류", "순위", "상호", "대표자", "소재지", "등록번호",
          "시공능력평가액", "공사실적평가액", "경영평가액", "기술능력평가액",
          "신인도평가액", "건설공사실적", "기술자수"]


def find_file(pattern, prefer_latest=False):
    """RAW_new에서 파일 찾기 (macOS NFD 파일명 대응: NFC로 정규화해 비교).

    후보가 여러 개면 --year(4자리, 2자리 모두)로 좁히고, 그래도 하나로 좁혀지지
    않으면 잘못된 연도 데이터를 조용히 쓰는 대신 후보를 보여주고 중단한다.

    prefer_latest: 파일명에 수집 시각이 붙는 소스(KICA_rank_YYYYMMDDHHMM)용.
    이런 파일명의 연도는 공시연도가 아니라 내려받은 날짜라 --year 매칭이 어긋나므로,
    연도로 좁히는 대신 파일명 역순 첫 번째(=가장 최근 수집분)를 쓴다."""
    cands = [Path(p) for p in sorted(glob.glob(str(RAW_NEW / "*")))
             if re.search(pattern, unicodedata.normalize("NFC", Path(p).name))]
    if not cands:
        raise SystemExit(f"오류: RAW_new에서 {pattern!r} 파일을 찾지 못함")
    if prefer_latest:
        return sorted(cands, key=lambda p: p.name, reverse=True)[0]
    if len(cands) > 1 and YEAR:
        yy = str(YEAR)[-2:]
        narrowed = [p for p in cands
                    if re.search(rf"(?<!\d){YEAR}(?!\d)|(?<!\d){yy}년", unicodedata.normalize("NFC", p.name))]
        if narrowed:
            cands = narrowed
    if len(cands) > 1:
        names = "\n    ".join(p.name for p in cands)
        raise SystemExit(
            f"오류: {pattern!r} 패턴에 파일이 {len(cands)}개 매칭됩니다. "
            f"--year로 연도를 지정하거나 대상 연도 파일만 남겨주세요:\n    {names}")
    return cands[0]


def fmt(v):
    """셀 값 → 문자열. float 정수(1687.0)는 int로, None은 빈 문자열."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def write_csv(name, rows):
    path = RAW_OUT / name
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(HEADER)
        w.writerows(rows)
    print(f"  ✔ raw/{name}: {len(rows):,}행")


def convert_cak():
    """종합건설 (대한건설협회). 시트 5개, 헤더 3~5행 병합, 데이터 6행부터.
    컬럼: 순위,상호,대표자,소재지,전화번호,등록번호,시공능력평가액,공사실적,경영,기술능력,신인도,
    [토건 시트: 토목분야액,건축분야액,실적(토목),실적(건축),기술자수] / [기타 시트: 건설공사실적,기술자수]"""
    sheet_to_sub = {"토건": "토목건축공사업", "토목": "토목공사업", "건축": "건축공사업",
                    "산설": "산업환경설비공사업", "조경": "조경공사업"}
    wb = openpyxl.load_workbook(find_file(r"^CAK_"), read_only=True)
    rows = []
    for sheet, sub in sheet_to_sub.items():
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=6, values_only=True):
            if not r[1] or not fmt(r[0]):   # 상호·순위 없는 행(빈행/각주) 제외
                continue
            is_togun = sheet == "토건"
            rows.append(["종합건설", sub, fmt(r[0]), fmt(r[1]), fmt(r[2]), fmt(r[3]),
                         fmt(r[5]),                        # 등록번호 (r[4]=전화번호는 표준 미포함)
                         fmt(r[6]), fmt(r[7]), fmt(r[8]), fmt(r[9]), fmt(r[10]),
                         "" if is_togun else fmt(r[11]),   # 건설공사실적 (토건은 토목/건축 분리라 비움)
                         fmt(r[15] if is_togun else r[12])])  # 기술자수
    wb.close()
    write_csv("general_construction.csv", rows)


def convert_specialty():
    """전문건설 (대한전문건설협회) '_업종' 파일. 업종별 12시트, 헤더 1~3행, 데이터 4행부터.
    컬럼: 연번(가나다순),상호,대표자,소재지,전화번호,등록번호,총액,공사실적,경영,기술능력,신인도,건설공사실적,기술자수,비고
    ※ '_주력분야' 파일은 별도 데이터셋으로 현재 미사용 (PRD §11)."""
    wb = openpyxl.load_workbook(find_file(r"전문건설.*_업종"), read_only=True)
    rows = []
    for ws in wb.worksheets:
        sub = ws.title.strip()
        for r in ws.iter_rows(min_row=4, values_only=True):
            if not r[1] or not fmt(r[0]):
                continue
            rows.append(["전문건설", sub, fmt(r[0]), fmt(r[1]), fmt(r[2]), fmt(r[3]),
                         fmt(r[5]), fmt(r[6]), fmt(r[7]), fmt(r[8]), fmt(r[9]),
                         fmt(r[10]), fmt(r[11]), fmt(r[12])])
    wb.close()
    write_csv("specialty_construction.csv", rows)


def convert_mechanical():
    """기계설비 (대한기계설비건설협회). 1시트, 헤더 3행, 데이터 4행부터. 원본에 '상호 가나다순' 명시.
    컬럼: 연번,상호,대표자명,소재지,전화번호,건설업등록번호,평가액,공사실적,경영,기술능력,신인도,건설공사실적,보유기술인수,비고"""
    wb = openpyxl.load_workbook(find_file(r"기계설비.*가스공사업"), read_only=True)
    ws = wb.worksheets[0]
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[1] or not fmt(r[0]):
            continue
        rows.append(["기계설비", "기계설비·가스공사업", fmt(r[0]), fmt(r[1]), fmt(r[2]), fmt(r[3]),
                     fmt(r[5]), fmt(r[6]), fmt(r[7]), fmt(r[8]), fmt(r[9]),
                     fmt(r[10]), fmt(r[11]), fmt(r[12])])
    wb.close()
    write_csv("mechanical_gas.csv", rows)


def convert_fire():
    """소방설비 (한국소방시설협회) CCE_*.xlsx. 헤더 3행, 데이터 4행부터. 단위: 천원(원본 명시).
    컬럼: 순번,구분,지역,상호,대표자,업종,등록번호,시공능력평가액,전국순위,지역순위
    표준 '순위'에는 전국순위 사용 (순번은 일련번호).
    ※ 전국순위는 업종(전문/일반(전기)/일반(기계))별로 따로 매겨지므로 중분류를 업종별로 분리한다.
    ※ 이 파일은 차원 메타데이터가 깨져 있어(1행으로 보고) reset_dimensions 필요."""
    sub_map = {"전문": "소방시설공사업(전문)",
               "일반(전기)": "소방시설공사업(일반·전기)",
               "일반(기계)": "소방시설공사업(일반·기계)"}
    wb = openpyxl.load_workbook(find_file(r"^CCE_"), read_only=True)
    ws = wb.worksheets[0]
    ws.reset_dimensions()
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r[3] or not fmt(r[8]):
            continue
        sub = sub_map.get(fmt(r[5]), f"소방시설공사업({fmt(r[5])})")
        rows.append(["소방설비", sub, fmt(r[8]), fmt(r[3]), fmt(r[4]), fmt(r[2]),
                     fmt(r[6]), fmt(r[7]), "", "", "", "", "", ""])
    wb.close()
    write_csv("fire_protection.csv", rows)


def convert_kica():
    """정보통신 (한국정보통신공사협회) KICA_rank_YYYYMMDDHHMM.csv. CP949 인코딩.
    컬럼: 순위,등록번호,상호,시공능력평가액(천원). 대표자·소재지 미제공.
    파일명 숫자는 공시연도가 아니라 수집 시각이므로 가장 최근 파일을 쓴다."""
    path = find_file(r"^KICA_rank_.*\.csv$", prefer_latest=True)
    print(f"    (정보통신 원본: {path.name})")
    rows = []
    with open(path, encoding="cp949", newline="") as f:
        for r in list(csv.reader(f))[1:]:
            if len(r) < 4 or not r[2].strip():
                continue
            rows.append(["정보통신", "정보통신공사업", r[0].strip(), r[2].strip(), "", "",
                         r[1].strip(), r[3].strip().replace(",", ""), "", "", "", "", "", ""])
    write_csv("ict_communication.csv", rows)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--year", help="대상 공시 연도 (예: 2026). 여러 연도 원본이 섞여 있을 때 사용")
    YEAR = ap.parse_args().year

    print(f"RAW_new 원본 → raw/ 표준 CSV 변환{f' (대상 연도: {YEAR})' if YEAR else ''}:")
    convert_cak()
    convert_specialty()
    convert_mechanical()
    convert_fire()
    convert_kica()
    print("완료. 다음 단계: python scripts/normalize.py")
